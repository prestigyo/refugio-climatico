#!/usr/bin/env python3
"""
Exporta a Excel los datos diarios completos (2017-2026) de una estación de
AEMET, con una hoja de resumen anual calculada CON FÓRMULAS (el libro se
recalcula solo si se corrigen datos). Para entregar a leads y prensa.

Lee   : aemet-temperaturas/datos/diarios_*.csv
Escribe: aemet-temperaturas/analisis/informes/datos-<slug>-2017-2026.xlsx

Uso:
    python scripts/exportar_excel_estacion.py --estacion 8293X
"""
from __future__ import annotations

import argparse
import csv
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import generar_calculadora as g

ARIAL = "Arial"
TEJA = "D9744E"


def cargar_diarios(indicativo: str) -> tuple[list[dict], str, str]:
    """Todas las filas diarias de la estación, ordenadas por fecha."""
    filas, nombre, prov = [], "", ""
    for ruta in sorted((g.AEMET_DIR / "datos").glob("diarios_2*.csv")):
        with ruta.open(newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                if r["indicativo"] != indicativo:
                    continue
                nombre, prov = r["nombre"], r["provincia"]

                def num(campo):
                    v = (r.get(campo) or "").strip()
                    try:
                        return float(v)
                    except ValueError:
                        return None
                filas.append({"fecha": r["fecha"], "tmin": num("tmin"),
                              "tmax": num("tmax"), "tmed": num("tmed"),
                              "prec": num("prec")})
    filas.sort(key=lambda x: x["fecha"])
    # el rolling (diarios_estaciones.csv) puede solapar con el año en curso:
    # deduplicado por fecha, la primera aparición gana
    vistas, unicas = set(), []
    for f in filas:
        if f["fecha"] in vistas:
            continue
        vistas.add(f["fecha"])
        unicas.append(f)
    return unicas, g.titular(nombre), g.PROVINCIAS.get(prov.strip().upper(),
                                                       g.titular(prov))


def construir_libro(ind: str, nombre: str, prov: str, filas: list[dict]) -> Workbook:
    wb = Workbook()

    # ---- Léeme -------------------------------------------------------------
    lee = wb.active
    lee.title = "Léeme"
    lee.column_dimensions["A"].width = 100
    textos = [
        (f"Datos diarios de la estación de AEMET {nombre} ({prov})", True),
        (f"Indicativo AEMET: {ind}", False),
        (f"Periodo: 2017 a 2026 · {len(filas)} días con registro", False),
        (f"Exportado el {g.fecha_es(date.today())} por nochetropical.es", False),
        ("", False),
        ("Hojas: «Datos diarios» (fecha, mínima, máxima, media, precipitación) y "
         "«Resumen anual» (calculado con fórmulas sobre los datos diarios: si "
         "corriges un dato, el resumen se recalcula).", False),
        ("Una NOCHE TROPICAL es aquella en que la temperatura mínima no baja de "
         "20 °C; una ECUATORIAL, de 25 °C. El verano es junio–agosto.", False),
        ("Las celdas vacías son días sin registro publicado por AEMET.", False),
        ("", False),
        ("Fuente: AEMET OpenData (https://opendata.aemet.es) · elaboración: "
         "nochetropical.es · datos bajo licencia CC BY 4.0 — puedes usarlos y "
         "publicarlos citando «Fuente: AEMET · nochetropical.es».", False),
    ]
    for i, (txt, negrita) in enumerate(textos, start=1):
        c = lee.cell(row=i, column=1, value=txt)
        c.font = Font(name=ARIAL, bold=negrita, size=12 if negrita else 10)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Datos diarios -----------------------------------------------------
    dd = wb.create_sheet("Datos diarios")
    cab = ["Fecha", "Mínima (°C)", "Máxima (°C)", "Media (°C)", "Precipitación (mm)"]
    for j, t in enumerate(cab, start=1):
        c = dd.cell(row=1, column=j, value=t)
        c.font = Font(name=ARIAL, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=TEJA)
    anchos = [12, 12, 12, 12, 16]
    for j, w in enumerate(anchos, start=1):
        dd.column_dimensions[chr(64 + j)].width = w
    for i, f in enumerate(filas, start=2):
        a, m, d = (int(x) for x in f["fecha"].split("-"))
        dd.cell(row=i, column=1, value=date(a, m, d)).number_format = "DD/MM/YYYY"
        for j, campo in enumerate(("tmin", "tmax", "tmed", "prec"), start=2):
            c = dd.cell(row=i, column=j, value=f[campo])
            c.number_format = "0.0"
    for fila in dd.iter_rows(min_row=2):
        for c in fila:
            c.font = Font(name=ARIAL, size=10)
    dd.freeze_panes = "A2"
    ult = len(filas) + 1

    # ---- Resumen anual (fórmulas sobre Datos diarios) ----------------------
    rs = wb.create_sheet("Resumen anual")
    cab2 = ["Año", "Noches tropicales (verano)", "Noches ecuatoriales (verano)",
            "Mínima media verano (°C)", "Máxima media verano (°C)",
            "Días con dato (verano)"]
    for j, t in enumerate(cab2, start=1):
        c = rs.cell(row=1, column=j, value=t)
        c.font = Font(name=ARIAL, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=TEJA)
        rs.column_dimensions[chr(64 + j)].width = 24
    anios = sorted({int(f["fecha"][:4]) for f in filas})
    hoy = date.today().year
    completos = [a for a in anios if a < hoy]
    F = f"'Datos diarios'!$A$2:$A${ult}"
    TN = f"'Datos diarios'!$B$2:$B${ult}"
    TX = f"'Datos diarios'!$C$2:$C${ult}"
    for i, anio in enumerate(anios, start=2):
        etiqueta = f"{anio} (en curso)" if anio >= hoy else str(anio)
        rs.cell(row=i, column=1, value=etiqueta).font = Font(name=ARIAL, size=10)
        ini, fin = f"DATE({anio},6,1)", f"DATE({anio},8,31)"
        rango = f'{F},">="&{ini},{F},"<="&{fin}'
        formulas = [
            f'=COUNTIFS({rango},{TN},">=20")',
            f'=COUNTIFS({rango},{TN},">=25")',
            f'=IFERROR(AVERAGEIFS({TN},{rango}),"")',
            f'=IFERROR(AVERAGEIFS({TX},{rango}),"")',
            f'=COUNTIFS({rango},{TN},"<>")',
        ]
        for j, fx in enumerate(formulas, start=2):
            c = rs.cell(row=i, column=j, value=fx)
            c.font = Font(name=ARIAL, size=10)
            c.number_format = "0.0" if j in (4, 5) else "0"
    # La media SOLO promedia veranos completos: el año en curso saldría bajo
    # (verano a medias) y contaminaría el resumen.
    fila_media = len(anios) + 2
    c = rs.cell(row=fila_media, column=1, value="Media (veranos completos)")
    c.font = Font(name=ARIAL, bold=True, size=10)
    for j, fmt in ((2, "0.0"), (3, "0.0"), (4, "0.0"), (5, "0.0")):
        col = chr(64 + j)
        c = rs.cell(row=fila_media, column=j,
                    value=f"=AVERAGE({col}2:{col}{len(completos) + 1})")
        c.font = Font(name=ARIAL, bold=True, size=10)
        c.number_format = fmt
    return wb


def exportar(ind: str, ruta_xlsx: Path) -> tuple[str, str, int]:
    """Genera el .xlsx en `ruta_xlsx`. Devuelve (nombre, provincia, n_días).
    Reutilizable desde generar_informe_lead.py."""
    filas, nombre, prov = cargar_diarios(ind)
    if not filas:
        raise SystemExit(f"Sin datos para {ind} en datos/diarios_*.csv")
    ruta_xlsx.parent.mkdir(parents=True, exist_ok=True)
    construir_libro(ind, nombre, prov, filas).save(ruta_xlsx)
    return nombre, prov, len(filas)


def main() -> int:
    ap = argparse.ArgumentParser(description="Exportar datos diarios de una estación a Excel")
    ap.add_argument("--estacion", required=True, help="indicativo AEMET (p. ej. 8293X)")
    args = ap.parse_args()
    ind = args.estacion.upper()
    filas, nombre, prov = cargar_diarios(ind)
    if not filas:
        raise SystemExit(f"Sin datos para {ind} en datos/diarios_*.csv")
    destino = g.AEMET_DIR / "analisis" / "informes"
    destino.mkdir(parents=True, exist_ok=True)
    ruta = destino / f"datos-{g.slug(nombre)}-2017-2026.xlsx"
    construir_libro(ind, nombre, prov, filas).save(ruta)
    print(f"OK -> {ruta}")
    print(f"   {nombre} ({prov}): {len(filas)} días")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
